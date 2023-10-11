import io
import os
import base64
import datetime
from PIL import Image
import streamlit as st
import pandas as pd
import re
import math
import plotly.graph_objects as go
from firebase_admin import db
from firebase_admin import credentials, storage
import openpyxl
from io import BytesIO
from formats import format_value, format_stat_value, indian_format
from streamlit_option_menu import option_menu


storage_bucket = os.getenv('STORAGE_BUCKET')


def show_profile(client_data):
    profile_picture = client_data.get("Profile Picture")
    # Display the profile picture if available
    if profile_picture is not None:
        # Decode base64 string to bytes
        profile_picture_bytes = base64.b64decode(profile_picture)

        # Convert profile picture from bytes to PIL Image
        image = Image.open(io.BytesIO(profile_picture_bytes))

       # Convert the image to RGB
        image_rgb = image.convert("RGB")

        # Save the image in JPG format with reduced quality (adjust the quality value as needed)
        image_path = "profile_picture.jpg"
        # Adjust the quality value as needed
        image_rgb.save(image_path, "JPEG", quality=50)

        # Define CSS style to position the profile picture in the right top corner with some margin
        css_style = f"""
            <style>
                .profile-picture-container {{
                    position: absolute;
                    top: -40px;  /* Adjust the top value as needed */
                    right: 10px;
                    border: 2px solid #ccc;
                    border-radius: 50%;
                    overflow: hidden;
                }}
                .profile-picture-container img {{
                    width: 100px;
                    height: 100px;
                }}
            </style>
        """

        # Display the CSS style
        st.markdown(css_style, unsafe_allow_html=True)

        # Display the profile picture in a container with the defined CSS style
        st.markdown(f"""
            <div class="profile-picture-container">
                <img src="data:image/jpeg;base64,{base64.b64encode(profile_picture_bytes).decode('utf-8')}" alt="Profile Picture">
            </div>
        """, unsafe_allow_html=True)

        # Remove the saved image file
        os.remove(image_path)

    pd.options.display.float_format = '{:,.2f}'.format
    # Set the title for the Streamlit app
    st.markdown("<h3 style='color: darkblue'>Profile</h3>",
                unsafe_allow_html=True)

    # Extract client data from the dictionary
    Name = client_data.get("Name", "")
    Username = client_data.get("Username", "")
    Email = client_data.get("Email", "")
    Password = client_data.get("Password", "")
    Phone_Number = client_data.get("Phone Number", "")
    Date_of_Birth = client_data.get("Date of Birth", "")
    Aadhar_Card_No = client_data.get("Aadhar Card No", "")
    PAN_Card_No = client_data.get("PAN Card No", "")
    Bank_Name = client_data.get("Bank Name", "")
    Bank_Account_No = client_data.get("Bank Account No", "")
    Brokers_list_1 = client_data.get("Brokers list 1", [])
    Brokers_list_2 = client_data.get("Brokers list 2", [])
    Strategy_list = client_data.get("Strategy list", [])
    Comments = client_data.get("Comments", "")
    Smart_Contract = client_data.get("Smart Contract", "")

    # Create a DataFrame to display the client data in tabular form
    data = {
        "Field": ["Name", "Username", "Email", "Password", "Phone Number", "Date of Birth", "Aadhar Card No",
                  "PAN Card No", "Bank Name", "Bank Account No", "Comments", "Smart Contract"],
        "Value": [str(Name), str(Username), str(Email), str(Password), str(Phone_Number), str(Date_of_Birth),
                  str(Aadhar_Card_No), str(PAN_Card_No), str(
                      Bank_Name), str(Bank_Account_No), str(Comments),
                  str(Smart_Contract)]
    }
    df = pd.DataFrame(data)
    # Display the DataFrame as a table with CSS styling and remove index column
    st.markdown(table_style, unsafe_allow_html=True)
    st.write(df.to_html(index=False, escape=False), unsafe_allow_html=True)

    # Display the broker list in vertical tabular form
    st.subheader("Brokers")
    st.write("Broker 1")
    if isinstance(Brokers_list_1, list) and len(Brokers_list_1) > 0:
        broker_1_data = {
            "Field": [],
            "Value": []
        }
        for broker_1 in Brokers_list_1:
            broker_1_data["Field"].extend(["Broker Name", "User Name", "Password", "2FA",
                                           "TotpAuth", "ApiCode", "ApiKey", "ApiSecret", "Active", "Capital", "Risk profile"])
            broker_1_data["Value"].extend([
                str(broker_1.get("broker_name", [""])[0]),
                str(broker_1.get("user_name", "")),
                str(broker_1.get("password", "")),
                str(broker_1.get("two_fa", "")),
                str(broker_1.get("totp_auth", "")),
                str(broker_1.get("api_code", "")),
                str(broker_1.get("api_key", "")),
                str(broker_1.get("api_secret", "")),
                str(broker_1.get("active", "")),
                "{:.2f}".format(broker_1.get("capital", 0.00)),
                str(broker_1.get("risk_profile", ""))
            ])

        broker_1_df = pd.DataFrame(broker_1_data)
        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(broker_1_df.to_html(index=False, escape=False),
                 unsafe_allow_html=True)

        # Add some space between the table and "Broker 2"
        st.markdown("<br>", unsafe_allow_html=True)  # Add this line
    else:
        st.warning("No broker data available.")

    st.write("Broker 2")
    if isinstance(Brokers_list_2, list) and len(Brokers_list_2) > 0:
        broker_2_data = {
            "Field": [],
            "Value": []
        }
        for broker_2 in Brokers_list_2:
            broker_2_data["Field"].extend(["Broker Name", "User Name", "Password", "2FA",
                                           "TotpAuth", "Apicode", "ApiKey", "ApiSecret", "Active", "Capital", "Risk profile"])
            broker_2_data["Value"].extend([
                str(broker_2.get("broker_name", [""])[0]),
                str(broker_2.get("user_name", "")),
                str(broker_2.get("password", "")),
                str(broker_2.get("two_fa", "")),
                str(broker_2.get("totp_auth", "")),
                str(broker_2.get("api_code", "")),
                str(broker_2.get("api_key", "")),
                str(broker_2.get("api_secret", "")),
                str(broker_2.get("active", "")),
                "{:.2f}".format(broker_2.get("capital", 0.00)),
                str(broker_2.get("risk_profile", ""))
            ])

        broker_2_df = pd.DataFrame(broker_2_data)
        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(broker_2_df.to_html(index=False, escape=False),
                 unsafe_allow_html=True)
    else:
        st.warning("No broker data available.")

     # Display the strategy list in vertical tabular form
    st.subheader("Strategies")
    if isinstance(Strategy_list, list) and len(Strategy_list) > 0:
        strategy_data = {
            "Strategy Name": [],
            "Broker": [],
            "Percentage Allocated": []
        }
        for strategy in Strategy_list:
            strategy_name = strategy.get("strategy_name", "")
            broker = strategy.get("broker", "")

            for selected_strategy in strategy_name:
                for selected_broker in broker:
                    perc_allocated_key = f"strategy_perc_allocated_{selected_strategy}_{selected_broker}_0"
                    percentage_allocated = strategy.get(perc_allocated_key, "")

                    strategy_data["Strategy Name"].append(selected_strategy)
                    strategy_data["Broker"].append(selected_broker)
                    strategy_data["Percentage Allocated"].append(
                        percentage_allocated)

        strategy_df = pd.DataFrame(strategy_data)
        # Display the DataFrame as a table with CSS styling and remove index column
        st.markdown(table_style, unsafe_allow_html=True)
        st.write(strategy_df.to_html(index=False, escape=False),
                 unsafe_allow_html=True)


table_style = """
<style>
table.dataframe {
    border-collapse: collapse;
    width: 100%;
}

table.dataframe th,
table.dataframe td {
    border: 1px solid black;
    padding: 8px;
    text-align: left; /* Align text to the left */
}

table.dataframe th {
    background-color: #f2f2f2;
}

table.dataframe tr:nth-child(even) {
    background-color: #f2f2f2;
}

table.dataframe tr:hover {
    background-color: #ddd;
}
</style>
"""

# Function to display performance dashboard


def display_performance_dashboard(selected_client, client_data, excel_file_name):
    # CSS style definitions for the option menu
    selected = option_menu(None, ["Calendar", "Statistics", "Graph"],
                           icons=['calendar', 'file-bar-graph', 'graph-up'],
                           menu_icon="cast", default_index=0, orientation="horizontal",
                           styles={
                               "container": {"padding": "0!important", "background-color": "#fafafa"},
                               "icon": {"color": "orange", "font-size": "25px"},
                               "nav-link": {"font-size": "25px", "text-align": "left", "margin": "0px", "--hover-color": "#eee"},
                               "nav-link-selected": {"background-color": "purple"},
    })

    # Reference the Firebase Storage bucket
    bucket = storage.bucket(storage_bucket)

    # Check if the client's Excel file exists in the Firebase Storage bucket
    blobs = bucket.list_blobs()
    file_exists = False
    for blob in blobs:
        if blob.name == excel_file_name:
            file_exists = True
            break

    data = []  # List to store extracted data from Excel
    # Excel header should be the same as below
    # SI NO,Date,Day,Transaction,Opening Balance,MP Wizard,AmiPy,ZRM,Overnight Options,
    # Gross PnL,Tax,Transaction Amount,Running Balance,Deposit/Withdrawal,Telegram Balance,Difference Amount,Remarks

    # If the client's Excel file exists, proceed to extract data
    if file_exists:
        # Reference the specific blob (file) in the bucket
        blob = bucket.blob(excel_file_name)

        # Download the blob into an in-memory bytes object
        byte_stream = BytesIO()
        blob.download_to_file(byte_stream)
        byte_stream.seek(0)

        # Load the Excel workbook from the bytes object
        wb = openpyxl.load_workbook(byte_stream, data_only=True)

        # Extract data if the "DTD" sheet exists in the workbook
        if "DTD" in wb.sheetnames:
            sheet = wb["DTD"]
            print(f"Extracting data from sheet: DTD")

            # Get column names and their indices from the first row
            column_indices = {cell.value: idx for idx,
                              cell in enumerate(sheet[1])}

            # Loop through each row in the sheet to read specific columns
            # Assuming headers are in the first row
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                date = row[column_indices['Date']].value
                day = row[column_indices['Day']].value
                trade_id = row[column_indices['Trade ID']].value
                details = row[column_indices['Details']].value
                amount = row[column_indices['Amount']].value
                running_balance = row[column_indices['Running Balance']].value

                data.append([date, day, trade_id, details,
                            amount, running_balance])

        # Add custom CSS for the table and value colors
        st.markdown("""
        <style>
        .custom-table {
            top: 3px;  /* Adjust the top value as needed */
            right: 500px;
            border: 2px solid #ccc;
            overflow: hidden;
            background-color: #E6E6FA;
            font-size: 19px;
            width: 100%;
            }
            .custom-table td {
            padding: 15px;  # Increase padding for larger cells
            border: 1px solid #ddd;  # Add borders to the cells
            }
        .positive-value {
            color: green;
        }
        .negative-value {
            color: red;
        }
        </style>
        """, unsafe_allow_html=True)

        # Calendar functionality
    if selected == "Calendar":
        selected_date = st.date_input("Select a Date")

        if selected_date:
            filtered_data = [record for record in data if record[0]
                             == selected_date.strftime('%Y-%m-%d')]

            # Debug print
            print(f"Filtered data for date {selected_date}: {filtered_data}")

            if filtered_data:
                table_data = []

                for record in filtered_data:
                    # Create a dictionary to store the labels and formatted values
                    field_names = {
                        # Index 3 corresponds to 'details'
                        "Details": format_value(record[3], "italic"),
                        # Index 4 corresponds to 'amount'
                        "Amount": format_value(record[4], "italic"),
                        # Index 5 corresponds to 'running_balance'
                        "Running Balance": format_value(record[5], "bold")
                    }

                    # Format the filtered data
                    formatted_data = [[field, value] for field,
                                      value in field_names.items() if value != "N/A"]

                    table_data.extend(formatted_data)

                # Display the table without header
                st.write(pd.DataFrame(table_data, columns=None).to_html(
                    classes='custom-table', header=False, index=False, escape=False), unsafe_allow_html=True)

    if selected == "Statistics":
        # Display date input fields for the user to select the start and end dates
        start_date = st.date_input(
            "Select Start Date", datetime.date(2023, 8, 4))
        end_date = st.date_input("Select End Date")

        # Filter the data based on the selected date range
        filtered_data = [record for record in data if record[0] is not None and start_date.strftime(
            '%Y-%m-%d') <= record[0] <= end_date.strftime('%Y-%m-%d')]

        # Extract relevant data from filtered_data
        opening_balances = [record[1] for record in filtered_data]
        running_balances = [record[10] for record in filtered_data]
        gross_pnls = [record[6] for record in filtered_data]
        net_pnls = [record[8] for record in filtered_data]
        transaction_amounts = [record[8] for record in filtered_data]
        deposit_withdrawals = [record[9] for record in filtered_data]
        tax = [record[7]for record in filtered_data]

        # Calculate statistics
        initial_capital = opening_balances[0] if opening_balances else 0
        ending_capital = running_balances[-1] if running_balances else 0
        total_profit = sum(gross_pnls)
        tax_amount = sum(tax)
        net_profit = total_profit - tax_amount
        net_profit_percent = (net_profit / initial_capital) * \
            100 if initial_capital != 0 else 0
        avg_profit = total_profit / len(gross_pnls) if gross_pnls else 0
        avg_profit_percent = (avg_profit / initial_capital) * \
            100 if initial_capital != 0 else 0
        total_deposits = sum(
            [amount for amount in deposit_withdrawals if amount is not None and amount > 0])
        total_withdrawal = sum(
            [amount for amount in deposit_withdrawals if amount is not None and amount < 0])
        total_commission = sum(transaction_amounts)

        # Create a DataFrame for the statistics
        stats_data = {
            "Metric": ["Initial Capital", "Ending Capital", "Total Profit", "Tax", "Net Profit", "Net Profit %",  "Avg. Profit", "Avg. Profit %",  "Total Deposits", "Total Withdrawal", "Total Commission"],
            "Value": [format_stat_value(initial_capital), format_stat_value(ending_capital), format_stat_value(total_profit), format_stat_value(tax_amount), format_stat_value(net_profit), format_stat_value(f"{net_profit_percent:.2f}%"),  format_stat_value(avg_profit), format_stat_value(f"{avg_profit_percent:.2f}%"),  format_stat_value(total_deposits), format_stat_value(total_withdrawal), format_stat_value(total_commission)]
        }

        stats_df = pd.DataFrame(stats_data)

        # Display the table without index and without column headers, and with custom styles
        st.write(stats_df.to_html(index=False, header=False,
                 classes='custom-table', escape=False), unsafe_allow_html=True)

    if selected == 'Graph':
        # If filtered_data is not defined, set it with a default date range
        try:
            filtered_data
        except NameError:
            start_date = datetime.date(2023, 8, 4)
            end_date = datetime.date.today()
            filtered_data = [record for record in data if record[0] is not None and start_date.strftime(
                '%Y-%m-%d') <= record[0] <= end_date.strftime('%Y-%m-%d')]

        graph_option = option_menu(None, ["Net PnL", "Running Balance"],
                                   # Assuming these are the icons you want
                                   icons=['line-chart', 'line-chart'],
                                   menu_icon="chart-bar",  # Placeholder icon
                                   default_index=0,
                                   orientation="horizontal",
                                   styles={
            "container": {"padding": "0!important", "background-color": "#fafafa"},
            "icon": {"color": "orange", "font-size": "18px"},
            "nav-link": {"font-size": "18px", "text-align": "left", "margin": "0px", "--hover-color": "#eee"},
            "nav-link-selected": {"background-color": "orange"},
        })

        if graph_option == "Net PnL":
            # Calculate net PnL for each record in filtered_data
            # Assuming gross_pnl - (transaction_amount/2) gives net PnL
            net_pnls = [record[6] - (record[8]/2) for record in filtered_data]

            # Create a Plotly figure
            fig = go.Figure()

            # Add traces for each segment of the line with the determined color
            for i in range(1, len(net_pnls)):
                color = 'green' if net_pnls[i] > net_pnls[i-1] else 'red'
                fig.add_trace(go.Scatter(x=[filtered_data[i-1][0], filtered_data[i][0]],
                                         y=[net_pnls[i-1], net_pnls[i]],
                                         mode='lines',
                                         line=dict(color=color, width=2),
                                         showlegend=False))  # Hide legend for each trace

            # Update the layout to hide the overall legend
            fig.update_layout(showlegend=False)

            # Display the graph using Streamlit's plotly_chart function
            st.plotly_chart(fig)

            def indian_format_decimal(num):
                """Format decimal number in Indian style with commas"""
                s = '{:,.2f}'.format(num)
                l, r = s.split('.')
                l = re.sub(r'(?<=\d)(?=(\d\d)+\d(?!\d))', ',', l)
                return l + '.' + r

        elif graph_option == "Running Balance":
            # Extract running balances from filtered_data
            running_balances = [record[10] for record in filtered_data]

            # Create a Plotly figure for Running Balance
            fig = go.Figure()

            # Add the running balances data to the figure
            fig.add_trace(go.Scatter(x=[record[0] for record in filtered_data],
                                     y=running_balances,
                                     mode='lines',
                                     line=dict(color='forestgreen', width=2),
                                     hovertemplate='%{y:,.2f}'))

            # Get the range of y-values for custom tick formatting
            y_max = max(running_balances)
            y_min = min(running_balances)
            tickvals = list(range(int(math.floor(y_min / 1e5) * 1e5),
                            int(math.ceil(y_max / 1e5) * 1e5), int(1e5)))
            ticktext = [indian_format(val) for val in tickvals]

            # Update y-axis to display values in Indian rupees with custom formatting
            fig.update_layout(
                yaxis_title="Amount (₹)",
                yaxis_tickvals=tickvals,
                yaxis_ticktext=ticktext
            )

            # Display the Running Balance graph using Streamlit's plotly_chart function
            st.plotly_chart(fig)
